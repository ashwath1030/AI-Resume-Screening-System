from openpyxl import Workbook
from openpyxl.styles import Font


def generate_excel_report(candidates, output_file):
    """
    Generate Candidate Ranking Excel Report
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidate Ranking"

    headers = [
        "Rank",
        "Candidate Name",
        "Email",
        "Skills",
        "Experience",
        "Score"
    ]

    # Header Row
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Candidate Data
    for row_num, candidate in enumerate(candidates, start=2):

        sheet.cell(row=row_num, column=1).value = row_num - 1
        sheet.cell(row=row_num, column=2).value = candidate["name"]
        sheet.cell(row=row_num, column=3).value = candidate["email"]
        sheet.cell(row=row_num, column=4).value = ", ".join(candidate["skills"])
        sheet.cell(row=row_num, column=5).value = candidate["experience"]
        sheet.cell(row=row_num, column=6).value = candidate["score"]

    workbook.save(output_file)